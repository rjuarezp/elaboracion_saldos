import pandas as pd
import os

def process_file(inputfile, outputpath, streamlit=False):
  try:

    if not streamlit:
      filename = inputfile.split(os.path.sep)[-1].split('.')[0].split('-')[0]
      df = pd.read_excel(inputfile, engine='openpyxl', skiprows=1)
      df.head()
    else:
      df = pd.read_excel(inputfile, engine='openpyxl', skiprows=1)
      filename = inputfile.name

    to_rename = [x for x in df.columns if x.startswith('Saldo')]
    df.rename(columns={to_rename[0]: 'Saldos', 'DESCRIPCION': ''}, inplace=True)

    to_drop = ['No vencido', 'Riesgo Máx.:', 'NIVEL']
    df.drop(to_drop, axis=1, inplace=True)

    cod_to_keep = ('80', '00', '04', '05')
    df = df[df['COD'].str.startswith(cod_to_keep)].copy()

    order = ['', 'COD', 'Saldos', 'Vencido', '1 - 30 días', '31 - 60 días', '61 - 90 días', '91 - 180 días', '181 - 365 días', '1 - 2 años', '+ 2 años']

    df_order = df[order].copy()

    df_order.reset_index(drop=True, inplace=True)

    df_order['index'] = df_order.index

    vendors = [x for x in df_order['COD'].unique() if x.startswith('80')]
    df_vendors = df_order[df_order['COD'].isin(vendors)].copy()
    n_vendors = len(vendors)
    mapping = {}
    for i in range(n_vendors):
      name = df_vendors.iloc[i, df_order.columns.get_loc("")]
      start = df_vendors.iloc[i, df_order.columns.get_loc("index")] + 1
      if i < n_vendors-1 :
        end = df_vendors.iloc[i+1, df_order.columns.get_loc("index")]
      else:
        end = df_order.shape[0]
      if end > start:
        mapping[vendors[i]] = {'start': start, 'end': end, 'name': name}

    columns = ['Vencido', '1 - 30 días', '31 - 60 días', '61 - 90 días', '91 - 180 días', '181 - 365 días', '1 - 2 años',
               '+ 2 años']
    cols_to_elminate = ['91 - 180 días', '181 - 365 días', '1 - 2 años', '+ 2 años']

    active_vendors = list(mapping.keys())
    total_df = pd.DataFrame()
    for ac_ve in active_vendors:
      vendor = df_order.iloc[mapping[ac_ve]['start']:mapping[ac_ve]['end']]
      mask = vendor['Vencido'] >= 0.5 # Sólo nos quedamos con los vencidos mayores de 0.5€
      vendor = vendor[mask].copy()
      vendor.sort_values(by=[''], inplace=True)
      line = pd.DataFrame({"": mapping[ac_ve]['name']}, index=[1])
      df1 = pd.concat([line, vendor.iloc[:]]).reset_index(drop=True)
      df1.drop('index', axis=1, inplace=True)
      new_line = {}
      for col in columns:
        new_line[col] = vendor[col].sum()
      to_drop = []
      for col in cols_to_elminate:
        if new_line[col] == 0:
          to_drop.append(col)
      line_end = pd.DataFrame(new_line, index=[1])
      df2 = pd.concat([df1, line_end]).reset_index(drop=True)
      df2.drop(to_drop, axis=1, inplace=True)
      total_df = pd.concat([total_df, df1])
      xlsx_filename = filename + '_' + mapping[ac_ve]['name'] + '.xlsx'
      outputname = os.path.join(outputpath, xlsx_filename)
      format_excel(df2, outputname)
    xlsx_filename_all = os.path.join(outputpath, filename + '_Fercampo.xlsx')
    total_df.to_excel(xlsx_filename_all, index=False)
    return 0
  except:
    return -1

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def auto_col_width(wrks):
  for col in wrks.columns:
     max_length = 0
     column = col[0].column # Get the column name
     # Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based) 
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = (max_length + 2) * 1.4
     wrks.column_dimensions[column].width = adjusted_width
     return wrks
 
def format_excel(data, name):
  wb = Workbook()
  ws = wb.active
  for r in dataframe_to_rows(data, index=False, header=True):
    ws.append(r)
  # Set auto width of columns
  for col in ws.columns:
    max_length = 0
    column = col[0].column # Get the column name
    # Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based) 
    for cell in col:
      try:  # Necessary to avoid error on empty cells
        if len(str(cell.value)) > max_length:
          max_length = len(cell.value)
      except:
        pass
    adjusted_width = (max_length + 2) * 1.1
    ws.column_dimensions[get_column_letter(column)].width = adjusted_width
  # Put columns in grey
  row_count = ws.max_row
  greyFill = PatternFill(start_color='00C0C0C0',
                   end_color='00C0C0C0',
                   fill_type='solid')
  for col in range(5,8):
    for row in range(1, row_count):
      cell = ws.cell(row=row, column=col)
      cell.fill = greyFill
  # Center the columns B to end
  row_count = ws.max_row
  column_count = ws.max_column
  for col in range(2, column_count+1):
    for row in range(1, row_count+1):
      cell = ws.cell(row=row, column=col)
      cell.alignment = Alignment(horizontal="center", vertical="center")
  # Apply format numbers
  for col in range(3, column_count+1):
    for row in range(3, row_count+1):
      cell = ws.cell(row=row, column=col)
      cell.number_format = '#,###;-#,###;;@'
  # Apply format
  ws.insert_rows(1)
  ws['A1'] = 'SALDO CLIENTES POR ANTIGÜEDAD'
  row_count = ws.max_row
  column_count = ws.max_column
  # Apply Arial font
  font = Font(name='Arial', size=10, color='FF000000')
  boldfont = Font(name='Arial', size=10, b=True, color='FF000000')
  for col in range(1, column_count+1):
    for row in range(1, row_count+1):
      cell = ws.cell(row=row, column=col)
      cell.font = font
  ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
  top_left_cell = ws['A1']
  top_left_cell.font  = boldfont
  top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
  for col in range(2, column_count + 1):
    cell = ws.cell(row=2, column=col)
    cell.font = boldfont
  seller_cell = ws['A3']
  seller_cell.font  = boldfont
  seller_cell.alignment = Alignment(horizontal="center", vertical="center")
  # Put the borders
  thin = Side(border_style="thin", color="000000")
  for col in range(1, column_count+1):
    for row in range(1, row_count):
      cell = ws.cell(row=row, column=col)
      cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
  # Saving document
  wb.save(name)

def read_v2022(filename):
  _data = pd.read_excel(filename, skiprows=3)
  _data = _data[:-1]  # Eliminar ultima fila
  _data['Cód. Cliente'] = _data['Cód. Cliente'].astype(int)
  _data['Cód. Cliente'] = _data['Cód. Cliente'].astype(str)
  _data['Cód. Cliente'] = _data['Cód. Cliente'].str.zfill(7)
  return _data


def process_file_v2022(inputfile, outputpath):
  try:
    df = read_v2022(inputfile)
    list_sellers = df['Vendedor'].unique()
    to_keep = ['Cliente', 'Cód. Cliente', 'Saldo Actual', 'Vencido', '1 - 30 días', '31 - 60 días', '61 - 90 días',
               '91 - 180 días', '181 - 365 días', '1 - 2 años', '+ 2 años']

    total = pd.DataFrame()
    for vendor in list_sellers:
      mask = df['Vendedor'] == vendor
      temp = df[mask][to_keep].copy()
      temp = temp[temp['Vencido'] >= 0.5].copy()  # Sólo nos quedamos con los vencidos mayores de 0.5€
      if len(temp) > 0:  # Si hay datos del vendedor...
        name_vendor = vendor.split('-')[-1].lstrip(' ').rstrip(' ')
        df_name = pd.DataFrame({'Cliente': name_vendor}, index=[1])
        temp1 = pd.concat([df_name, temp]).reset_index(drop=True)
        temp1.rename(columns={'Cliente': ''}, inplace=True)
        cols = ['Vencido', '1 - 30 días', '31 - 60 días', '61 - 90 días', '91 - 180 días', '181 - 365 días', '1 - 2 años',
                '+ 2 años']
        last_row = dict()
        for col in cols:
          if temp1[col].sum() == 0:
            last_row[col] = ''
          else:
            last_row[col] = temp1[col].sum()
        lastrow_df = pd.DataFrame(last_row, index=[0])
        temp2 = pd.concat([temp1, lastrow_df]).reset_index(drop=True)

        #print('Processing {}'.format(vendor))
        outputname = os.path.join(outputpath, 'data_' + name_vendor.replace(' ', '_') + '.xlsx')
        # temp.to_excel(outputname, index=False)
        format_excel(temp2, outputname)
        total = pd.concat([total, temp1], ignore_index=True).reset_index(drop=True)
    outputname_total = os.path.join(outputpath, 'Fercampo_output.xlsx')
    total.to_excel(outputname_total, index=False)
    return 0
  except:
    return -1